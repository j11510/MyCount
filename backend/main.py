from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from fastapi.middleware.cors import CORSMiddleware
from datetime import timedelta
from typing import List, Optional

import models, schemas, crud, auth
from database import engine, get_db

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="MyCount API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow all origins for dev
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/token", response_model=schemas.Token)
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    admin = auth.authenticate_admin(db, form_data.username, form_data.password)
    if not admin:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": admin.username, "role": admin.role}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users", response_model=List[schemas.AdminResponse])
def read_users(db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.get_admins(db)

@app.get("/me", response_model=schemas.AdminResponse)
def read_me(current_user: models.Admin = Depends(auth.get_current_user)):
    return current_user

@app.post("/users", response_model=schemas.AdminResponse)
def create_user(user: schemas.AdminCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    db_user = crud.get_admin_by_username(db, username=user.username)
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    return crud.create_admin(db=db, admin=user)

@app.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    # Prevent self-deletion
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    crud.delete_admin(db, user_id)
    return {"status": "deleted"}

@app.get("/fixed-expenses", response_model=List[schemas.FixedExpenseResponse])
def read_fixed_expenses(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.get_fixed_expenses(db, skip=skip, limit=limit)

@app.post("/fixed-expenses", response_model=schemas.FixedExpenseResponse)
def create_fixed_expense(expense: schemas.FixedExpenseCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.create_fixed_expense(db=db, expense=expense)

@app.delete("/fixed-expenses/{expense_id}")
def delete_fixed_expense(expense_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    crud.delete_fixed_expense(db, expense_id)
    return {"status": "deleted"}

@app.post("/monthly-records", response_model=schemas.MonthlyRecordResponse)
def create_monthly_record(record: schemas.MonthlyRecordCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    db_record = crud.get_monthly_record_by_date(db, year=record.year, month=record.month)
    if db_record:
        raise HTTPException(status_code=400, detail="Record for this month already exists")
    return crud.create_monthly_record(db=db, record=record)

@app.get("/monthly-records", response_model=List[schemas.MonthlyRecordResponse])
def read_monthly_records(skip: int = 0, limit: int = 12, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.get_monthly_records(db, skip=skip, limit=limit)

@app.get("/monthly-records/{record_id}", response_model=schemas.MonthlyRecordResponse)
def read_monthly_record(record_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    db_record = crud.get_monthly_record(db, record_id=record_id)
    if db_record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    return db_record

@app.put("/monthly-records/{record_id}/balance", response_model=schemas.MonthlyRecordResponse)
def update_balance(record_id: int, current_balance: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.update_monthly_record_balance(db, record_id, current_balance)

@app.delete("/monthly-records/{record_id}")
def delete_monthly_record(record_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    crud.delete_monthly_record(db, record_id)
    return {"status": "deleted"}

@app.post("/monthly-records/{record_id}/items", response_model=schemas.MonthlyItemResponse)
def create_item_for_record(record_id: int, item: schemas.MonthlyItemCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.create_monthly_item(db=db, record_id=record_id, item=item)

@app.put("/monthly-items/{item_id}", response_model=schemas.MonthlyItemResponse)
def update_item_amount(item_id: int, amount: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    return crud.update_monthly_item(db=db, item_id=item_id, amount=amount)

@app.delete("/monthly-items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_admin)):
    crud.delete_monthly_item(db=db, item_id=item_id)
    return {"status": "deleted"}

# Accounting Modules
@app.get("/accounting/categories", response_model=List[schemas.AccountingCategoryResponse])
def read_accounting_categories(type: Optional[str] = None, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.get_accounting_categories(db, type=type)

@app.post("/accounting/categories", response_model=schemas.AccountingCategoryResponse)
def create_accounting_category(category: schemas.AccountingCategoryCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.create_accounting_category(db, category)

@app.delete("/accounting/categories/{category_id}")
def delete_accounting_category(category_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    crud.delete_accounting_category(db, category_id)
    return {"status": "deleted"}

@app.get("/accounting/transactions/{bank_account}", response_model=List[schemas.AccountingRecordResponse])
def read_accounting_transactions(bank_account: str, year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.get_accounting_records(db, bank_account, year, month)

@app.post("/accounting/transactions", response_model=schemas.AccountingRecordResponse)
def create_accounting_transaction(record: schemas.AccountingRecordCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.create_accounting_record(db, record)

@app.delete("/accounting/transactions/{record_id}")
def delete_accounting_transaction(record_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    crud.delete_accounting_record(db, record_id)
    return {"status": "deleted"}

@app.put("/accounting/transactions/{record_id}", response_model=schemas.AccountingRecordResponse)
def update_accounting_transaction(record_id: int, record: schemas.AccountingRecordCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    db_record = crud.update_accounting_record(db, record_id, record)
    if not db_record:
        raise HTTPException(status_code=404, detail="Record not found")
    return db_record

@app.get("/accounting/stats")
def read_accounting_stats(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    stats = crud.get_accounting_stats(db, year, month)
    result = []
    for s in stats:
        result.append({
            "bank_account": s[0],
            "category_name": s[1],
            "type": s[2],
            "amount": s[3]
        })
    return result

@app.get("/accounting/balances")
def read_accounting_balances(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    import calendar
    prev_year, prev_month = (year, month - 1) if month > 1 else (year - 1, 12)
    
    results = {}
    for code in ["finances", "donations", "meeting"]:
        opening = crud.get_account_running_balance(db, code, prev_year, prev_month)
        closing = crud.get_account_running_balance(db, code, year, month)
        results[code] = {
            "opening": opening,
            "closing": closing
        }
    return results

@app.get("/accounting/accounts", response_model=List[schemas.AccountingAccountResponse])
def read_accounting_accounts(db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.get_accounting_accounts(db)

@app.put("/accounting/accounts/{code}/balance", response_model=schemas.AccountingAccountResponse)
def update_account_balance(code: str, balance: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.update_accounting_account_balance(db, code, balance)
@app.get("/donations", response_model=List[schemas.DonationRecordResponse])
def read_donations(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.get_donation_records(db, year, month)

@app.post("/donations", response_model=schemas.DonationRecordResponse)
def create_donation(record: schemas.DonationRecordCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.create_donation_record(db, record)

@app.delete("/donations/{record_id}")
def delete_donation(record_id: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.delete_donation_record(db, record_id)

@app.put("/donations/{record_id}", response_model=schemas.DonationRecordResponse)
def update_donation(record_id: int, record: schemas.DonationRecordCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    db_record = crud.update_donation_record(db, record_id, record)
    if not db_record:
        raise HTTPException(status_code=404, detail="Record not found")
    return db_record

@app.get("/donations/donors")
def read_donors(db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    donors = crud.get_unique_donors(db)
    return [d[0] for d in donors]

# Monthly Reports Endpoints
@app.get("/monthly-reports", response_model=Optional[schemas.MonthlyReportResponse])
def read_monthly_report(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.get_monthly_report(db, year, month)

@app.post("/monthly-reports", response_model=schemas.MonthlyReportResponse)
def create_monthly_report(report: schemas.MonthlyReportCreate, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.upsert_monthly_report(db, report)

@app.get("/monthly-reports/export")
def export_monthly_report(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    report = crud.get_monthly_report(db, year, month)
    settlement = crud.get_infant_expenses_from_ledger(db, year, month)
    
    import io
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}.{month:02d} 보고"
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 1. Main Title
    ws.merge_cells('A1:J1')
    ws['A1'] = f"영아부 보고 및 계획(안)"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    # Reporter Info
    reporter_name = report.reporter if report else ""
    ws.merge_cells('A2:J2')
    ws['A2'] = f"◆ 보고자 : {reporter_name} / ◆ {month}월 사역보고 ({year}.{month:02d})"
    ws['A2'].alignment = Alignment(horizontal='right')
    
    # Section A & B: Business Report and Next Month Plan (2 columns layout)
    ws.merge_cells('A3:E3')
    ws['A3'] = f"{month}월 사업보고 (결산)"
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = header_fill
    ws['A3'].border = thin_border
    ws['A3'].alignment = center_align
    
    ws.merge_cells('F3:J3')
    ws['F3'] = f"{month+1 if month < 12 else 1}월 사업계획 (안)"
    ws['F3'].font = Font(bold=True)
    ws['F3'].fill = header_fill
    ws['F3'].border = thin_border
    ws['F3'].alignment = center_align
    
    # Headers for A & B
    headers = ['항목', '예산', '결산', '항목', '예산']
    # A(A-C), B(D-E) - wait HWP was grouped.
    # Let's use A-C for Report, D-E for Plan. Or just follow the HWP exactly.
    # HWP had A: Report(Item, Budget, Settlement) and B: Plan(Item, Budget)
    
    ws.merge_cells('A4:B4')
    ws['A4'] = "사업 항목"
    ws['C4'] = "예산"
    ws['D4'] = "결산"
    ws.merge_cells('F4:I4')
    ws['F4'] = "사업 항목"
    ws['J4'] = "예산"
    
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
         ws[f'{c}4'].border = thin_border
         ws[f'{c}4'].fill = header_fill
         ws[f'{c}4'].alignment = center_align

    # Fill rows for Report and Plan
    plan_list = json.loads(report.plan_data if report else "[]")
    max_len = max(len(settlement), len(plan_list), 5)
    
    curr_row = 5
    for i in range(max_len):
        r_idx = curr_row + i
        # Report side (Left)
        ws.merge_cells(f'A{r_idx}:B{r_idx}')
        if i < len(settlement):
            ws[f'A{r_idx}'] = settlement[i].description
            ws[f'D{r_idx}'] = settlement[i].amount
        
        # Plan side (Right)
        ws.merge_cells(f'F{r_idx}:I{r_idx}')
        if i < len(plan_list):
            ws[f'F{r_idx}'] = plan_list[i].get('item', '')
            ws[f'J{r_idx}'] = plan_list[i].get('budget', 0)
            
        for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{c}{r_idx}'].border = thin_border
            
    # Sum rows
    sum_row = curr_row + max_len
    ws.merge_cells(f'A{sum_row}:C{sum_row}')
    ws[f'A{sum_row}'] = "계 (총합)"
    ws[f'D{sum_row}'] = sum(e.amount for e in settlement)
    
    ws.merge_cells(f'F{sum_row}:I{sum_row}')
    ws[f'F{sum_row}'] = "계 (총합)"
    ws[f'J{sum_row}'] = sum(p.get('budget', 0) for p in plan_list)
    
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws[f'{c}{sum_row}'].border = thin_border
        ws[f'{c}{sum_row}'].font = Font(bold=True)

    # Section C: Attendance
    start_att = sum_row + 2
    ws.merge_cells(f'A{start_att}:J{start_att}')
    ws[f'A{start_att}'] = "출석 현황 및 헌금 내역"
    ws[f'A{start_att}'].font = Font(bold=True)
    ws[f'A{start_att}'].fill = header_fill
    ws[f'A{start_att}'].border = thin_border
    ws[f'A{start_att}'].alignment = center_align
    
    att_header = start_att + 1
    ws[f'A{att_header}'] = "요일(주차)"
    ws.merge_cells(f'B{att_header}:D{att_header}')
    ws[f'B{att_header}'] = "영아 출석"
    ws.merge_cells(f'E{att_header}:G{att_header}')
    ws[f'E{att_header}'] = "교사 출석"
    ws.merge_cells(f'H{att_header}:I{att_header}')
    ws[f'H{att_header}'] = "헌금"
    ws[f'J{att_header}'] = "비고"
    
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws[f'{c}{att_header}'].border = thin_border
        ws[f'{c}{att_header}'].fill = header_fill
        ws[f'{c}{att_header}'].alignment = center_align
        
    att_list = json.loads(report.attendance_data if report else "[]")
    for i, a in enumerate(att_list):
        r_idx = att_header + 1 + i
        ws[f'A{r_idx}'] = a.get('date', '')
        ws.merge_cells(f'B{r_idx}:D{r_idx}')
        ws[f'B{r_idx}'] = a.get('kids', 0)
        ws.merge_cells(f'E{r_idx}:G{r_idx}')
        ws[f'E{r_idx}'] = a.get('teachers', 0)
        ws.merge_cells(f'H{r_idx}:I{r_idx}')
        ws[f'H{r_idx}'] = f"{a.get('donation', 0):,}원"
        ws[f'J{r_idx}'] = a.get('note', '')
        
        for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{c}{r_idx}'].border = thin_border

    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=monthly_report_{year}_{month}.xlsx"})

# Infant Expenses (from Ledger)
@app.get("/infant-expenses", response_model=List[schemas.AccountingRecordResponse])
def read_infant_expenses(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    return crud.get_infant_expenses_from_ledger(db, year, month)

# We remove POST/PUT/DELETE for infant-expenses as they should be managed via Ledger
@app.get("/infant-expenses/export")
def export_infant_expenses(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    expenses = crud.get_infant_expenses_from_ledger(db, year, month)
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}.{month:02d}"
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    ws.merge_cells('F2:M2')
    ws['F2'] = f"{year}-{month:02d}-24 기준"
    ws['F2'].alignment = Alignment(horizontal='right')
    
    ws.merge_cells('B3:F3')
    ws['B3'] = f"영아부 지출내역 ({year}년 {month}월)"
    ws['B3'].font = Font(bold=True, size=12)
    ws['B3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('H3:M3')
    ws['H3'] = f"지출 세부내역 ({year}년 {month}월)"
    ws['H3'].font = Font(bold=True, size=12)
    ws['H3'].alignment = Alignment(horizontal='center')
    
    headers_left = ['NO', '항목', '금액', '결제방법', '비고']
    headers_right = ['NO', '항목', '금액', '세부 산정 금액', '결제방법', '비고']
    
    for idx, h in enumerate(headers_left):
        cell = ws.cell(row=4, column=idx+2)
        cell.value = h
        cell.border = thin_border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    for idx, h in enumerate(headers_right):
        cell = ws.cell(row=4, column=idx+8)
        cell.value = h
        cell.border = thin_border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    curr_row = 5
    total_sum = sum(e.amount for e in expenses)
    
    for idx, e in enumerate(expenses):
        row_idx = curr_row + idx
        no = idx + 1
        
        ws.cell(row=row_idx, column=2, value=no).border = thin_border
        ws.cell(row=row_idx, column=3, value=e.description).border = thin_border
        ws.cell(row=row_idx, column=4, value=e.amount).border = thin_border
        ws.cell(row=row_idx, column=5, value=e.bank_account).border = thin_border
        ws.cell(row=row_idx, column=6, value=e.date.strftime('%m/%d')).border = thin_border
        
        ws.cell(row=row_idx, column=8, value=no).border = thin_border
        ws.cell(row=row_idx, column=9, value=e.description).border = thin_border
        ws.cell(row=row_idx, column=10, value=e.amount).border = thin_border
        ws.cell(row=row_idx, column=11, value=None).border = thin_border
        ws.cell(row=row_idx, column=12, value=e.bank_account).border = thin_border
        ws.cell(row=row_idx, column=13, value=e.date.strftime('%m/%d')).border = thin_border

    sum_row = curr_row + len(expenses)
    ws.cell(row=sum_row, column=2, value="합계").border = thin_border
    ws.cell(row=sum_row, column=3).border = thin_border
    ws.cell(row=sum_row, column=4, value=total_sum).border = thin_border
    ws.cell(row=sum_row, column=5).border = thin_border
    ws.cell(row=sum_row, column=6).border = thin_border
    
    ws.cell(row=sum_row, column=2).font = Font(bold=True)
    ws.cell(row=sum_row, column=4).font = Font(bold=True)

    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=infant_expense_{year}_{month}.xlsx"})

@app.get("/accounting/export")
def export_accounting_records(bank_account: str, year: Optional[int] = None, month: Optional[int] = None, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    
    records = crud.get_accounting_records(db, bank_account, year, month)
    # Join with category to get names
    data = []
    for r in records:
        cat = db.query(models.AccountingCategory).filter(models.AccountingCategory.id == r.category_id).first()
        data.append({
            "날짜": r.date,
            "분류": cat.name if cat else "미분류",
            "내역": r.description,
            "수입": r.amount if r.type == 'income' else 0,
            "지출": r.amount if r.type == 'expense' else 0,
            "비고": r.remarks or ""
        })
    
    wb = openpyxl.Workbook()
    ws = wb.active
    account_name = next((a["name"] for a in [{"id":"finances","name":"재정 통장"},{"id":"donations","name":"찬조금 통장"},{"id":"meeting","name":"모임 통장"}] if a["id"] == bank_account), bank_account)
    title = f"{account_name} 거래 내역"
    if year and month:
        title += f" ({year}년 {month}월)"
    else:
        title += " (전체)"
    
    ws.title = "거래내역"
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ["날짜", "분류", "내역", "수입", "지출", "비고"]
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    for r_idx, row_data in enumerate(data, 4):
        ws.cell(row=r_idx, column=1, value=row_data["날짜"]).border = thin_border
        ws.cell(row=r_idx, column=2, value=row_data["분류"]).border = thin_border
        ws.cell(row=r_idx, column=3, value=row_data["내역"]).border = thin_border
        ws.cell(row=r_idx, column=4, value=row_data["수입"]).border = thin_border
        ws.cell(row=r_idx, column=5, value=row_data["지출"]).border = thin_border
        ws.cell(row=r_idx, column=6, value=row_data["비고"]).border = thin_border
        
        # Format numbers
        ws.cell(row=r_idx, column=4).number_format = '#,##0'
        ws.cell(row=r_idx, column=5).number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ledger_{bank_account}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/donations/export")
def export_donations(year: int, month: int, db: Session = Depends(get_db), current_user: models.Admin = Depends(auth.check_manager)):
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    records = crud.get_donation_records(db, year, month)
    data = []
    total = 0
    current_date = None
    day_total = 0
    for r in records:
        if current_date and r.date != current_date:
            data.append({"날짜": f"{current_date} 소계", "성함": "-", "금액": day_total})
            day_total = 0
        data.append({"날짜": r.date, "성함": r.member_name, "금액": r.amount})
        current_date = r.date
        day_total += r.amount
        total += r.amount
    
    if current_date:
        data.append({"날짜": f"{current_date} 소계", "성함": "-", "금액": day_total})
    if data:
        data.append({"날짜": "월 합계", "성함": "", "금액": total})
        
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='헌금내역')
    output.seek(0)
    
    headers = {'Content-Disposition': f'attachment; filename="donation_{year}_{month}.xlsx"'}
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
